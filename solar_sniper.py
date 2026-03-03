import asyncio
import io
import json
import os
import re
import time
import uuid
from dataclasses import dataclass, field
from typing import Any, AsyncGenerator, Dict, List, Literal, Optional, Tuple

from fastapi import BackgroundTasks, FastAPI, HTTPException
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from pydantic import BaseModel, Field

try:
    from playwright.sync_api import sync_playwright
except Exception:
    sync_playwright = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception:
    Workbook = None


APP_TITLE = "SOLAR SNIPER - Energy Lead Extractor"


class StartJobRequest(BaseModel):
    category: str = Field(min_length=2, max_length=200)
    city: str = Field(min_length=2, max_length=120)


class JobStatus(BaseModel):
    id: str
    state: Literal["queued", "running", "done", "error"]
    progress: int
    message: str
    started_at: float
    finished_at: Optional[float] = None
    error: Optional[str] = None
    results_count: int = 0


class LeadRow(BaseModel):
    business_name: str
    phone: str
    address: str
    category: str
    email: str


@dataclass
class Job:
    id: str
    category: str
    city: str
    state: str = "queued"
    progress: int = 0
    message: str = "Queued"
    started_at: float = field(default_factory=lambda: time.time())
    finished_at: Optional[float] = None
    error: Optional[str] = None
    results: List[LeadRow] = field(default_factory=list)
    events: asyncio.Queue = field(default_factory=asyncio.Queue)

    async def emit(self, progress: int, message: str) -> None:
        self.progress = max(0, min(100, int(progress)))
        self.message = message
        await self.events.put(
            {
                "progress": self.progress,
                "message": self.message,
                "state": self.state,
                "error": self.error,
                "results_count": len(self.results),
            }
        )


app = FastAPI(title=APP_TITLE, version="1.0.0")
JOBS: Dict[str, Job] = {}


def _normalize_phone_digits(phone: str) -> str:
    p = (phone or "").strip()
    p = p.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    p = re.sub(r"\D+", "", p)
    if p.startswith("0039"):
        p = p[4:]
    if p.startswith("39") and len(p) > 10:
        p = p[2:]
    return p


def clean_phone(raw_phone):
    if not raw_phone:
        return None, False

    num = re.sub(r"[^\d+]", "", str(raw_phone))

    if num.startswith("+39"):
        num = num[3:]
    elif num.startswith("0039"):
        num = num[4:]

    if not num:
        return None, False

    if num.startswith("0"):
        if num.startswith("00"):
            num = num[1:]
        return num, False
    elif num.startswith("3"):
        return num, True
    return num, False


def extract_phone_from_html(soup):
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"].lower()
        if "tel:" in href or "wa.me/" in href or "api.whatsapp.com/send" in href:
            raw_number = (
                href.replace("tel:", "")
                .replace("https://wa.me/", "")
                .replace("http://wa.me/", "")
                .split("?")[0]
            )
            num = re.sub(r"[^\d+]", "", raw_number)
            if num:
                return num
    try:
        text = soup.get_text(" ", strip=True)
    except Exception:
        text = ""
    m = re.search(r"\b(3[2-9]\d{1}\s?\d{3}\s?\d{3,4})\b", text or "")
    if m:
        cand = m.group(1)
        if cand and not str(cand).startswith("0"):
            return cand
    return None


def _is_mobile(phone: str) -> bool:
    d, is_m = clean_phone(phone)
    if not d or not is_m:
        return False
    return bool(re.fullmatch(r"3[2-9]\d{8}", str(d)))


def _phone_display(phone: str) -> str:
    raw = (phone or "").strip()
    if not raw:
        return ""
    if _is_mobile(raw):
        return f"📱 {raw}"
    return raw


def _sort_key(row: LeadRow) -> Tuple[int, str]:
    pri = 0 if _is_mobile(row.phone) else 1
    return (pri, row.business_name.lower())


def _extract_email_from_html(html: str) -> str:
    if not html:
        return ""
    m = re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", html, flags=re.IGNORECASE)
    if not m:
        return ""
    for e in m:
        if "noreply" in e.lower() or "no-reply" in e.lower():
            continue
        return e
    return m[0]


def _fetch_website_email(website: str) -> str:
    if not website:
        return ""
    try:
        import requests

        r = requests.get(website, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code >= 400:
            return ""
        return _extract_email_from_html(r.text or "")
    except Exception:
        return ""


def _scrape_google_maps_sync(category: str, city: str) -> List[Dict[str, Any]]:
    if sync_playwright is None:
        raise RuntimeError("Playwright non installato")

    query = f"{category} {city}".strip()
    results: List[Dict[str, Any]] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(
            channel="chrome",
            headless=False,
            args=[
                "--lang=it-IT",
                "--disable-blink-features=AutomationControlled",
                "--no-default-browser-check",
            ],
        )
        context = browser.new_context(
            locale="it-IT",
            timezone_id="Europe/Rome",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1400, "height": 900},
        )
        page = context.new_page()
        page.set_default_timeout(25000)

        search_url = "https://www.google.com/maps/search/" + re.sub(r"\s+", "%20", query) + "?hl=it&gl=it"
        page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(1500)

        try:
            if page.locator('#L2AGLb').count() and page.locator('#L2AGLb').first.is_visible():
                page.locator('#L2AGLb').first.click(timeout=2500)
                page.wait_for_timeout(800)
        except Exception:
            pass

        cards = page.locator('div[role="article"]')
        alt_cards = page.locator('div.Nv2PK')

        for _ in range(25):
            if cards.count() > 0 or alt_cards.count() > 0:
                break
            page.wait_for_timeout(800)

        card_list = cards if cards.count() > 0 else alt_cards
        max_cards = min(40, card_list.count())

        for i in range(max_cards):
            try:
                c = card_list.nth(i)
                if not c.is_visible():
                    continue
                c.click(timeout=3000)
                page.wait_for_timeout(1100)

                name = ""
                try:
                    name = page.locator('h1.DUwDvf').first.inner_text(timeout=2500).strip()
                except Exception:
                    name = ""

                address = ""
                try:
                    addr_btn = page.locator('button[data-item-id="address"]').first
                    if addr_btn.count():
                        address = addr_btn.inner_text(timeout=2000).strip()
                except Exception:
                    address = ""

                phone = ""
                try:
                    phone_btn = page.locator('button[data-item-id^="phone:"]').first
                    if phone_btn.count():
                        phone = phone_btn.inner_text(timeout=2000).strip()
                except Exception:
                    phone = ""

                website = ""
                try:
                    web_a = page.locator('a[data-item-id="authority"]').first
                    if web_a.count():
                        website = web_a.get_attribute("href") or ""
                except Exception:
                    website = ""

                if not name:
                    continue

                results.append(
                    {
                        "business_name": name,
                        "address": address,
                        "phone": phone,
                        "website": website,
                    }
                )
            except Exception:
                continue

        try:
            context.close()
        except Exception:
            pass
        try:
            browser.close()
        except Exception:
            pass

    return results


async def run_job(job: Job) -> None:
    job.state = "running"
    await job.emit(5, "Ricerca su Google Maps...")

    try:
        raw = await asyncio.to_thread(_scrape_google_maps_sync, job.category, job.city)
        if not raw:
            raise RuntimeError("Nessun risultato trovato.")

        await job.emit(25, f"Trovate {len(raw)} attività. Estrazione contatti...")

        leads: List[LeadRow] = []
        for idx, r in enumerate(raw, start=1):
            phone = str(r.get("phone") or "").strip()
            if not phone:
                continue

            email = ""
            website = str(r.get("website") or "").strip()
            if website:
                email = await asyncio.to_thread(_fetch_website_email, website)
                try:
                    import requests

                    rr = requests.get(website, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
                    if int(getattr(rr, "status_code", 0) or 0) < 400:
                        html = rr.text or ""
                        if html.strip():
                            from bs4 import BeautifulSoup

                            soup = BeautifulSoup(html, "html.parser")
                            raw_site_phone = extract_phone_from_html(soup)
                            cand, cand_is_mobile = clean_phone(raw_site_phone)
                            if cand and cand_is_mobile and re.fullmatch(r"3[2-9]\d{8}", str(cand)):
                                phone = cand
                except Exception:
                    pass

            leads.append(
                LeadRow(
                    business_name=str(r.get("business_name") or "").strip(),
                    phone=phone,
                    address=str(r.get("address") or "").strip(),
                    category=job.category,
                    email=email,
                )
            )

            if idx % 8 == 0:
                await job.emit(25 + min(60, int((idx / max(1, len(raw))) * 60)), "Estrazione contatti...")

        leads_sorted = sorted(leads, key=_sort_key)
        cleaned: List[LeadRow] = []
        for x in leads_sorted:
            cleaned.append(
                LeadRow(
                    business_name=x.business_name,
                    phone=_phone_display(x.phone),
                    address=x.address,
                    category=x.category,
                    email=x.email,
                )
            )

        job.results = cleaned
        job.state = "done"
        job.finished_at = time.time()
        await job.emit(100, f"Completato: {len(job.results)} lead pronti (solo con telefono).")
    except Exception as e:
        job.state = "error"
        job.finished_at = time.time()
        job.error = str(e)
        await job.emit(100, f"Errore: {job.error}")


@app.get("/", response_class=HTMLResponse)
async def index() -> str:
    return HTMLResponse(_INDEX_HTML.replace("__APP_TITLE__", APP_TITLE))


@app.get("/health")
async def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.post("/jobs", response_model=JobStatus)
async def start_job(payload: StartJobRequest, background: BackgroundTasks) -> JobStatus:
    job_id = str(uuid.uuid4())
    job = Job(id=job_id, category=payload.category.strip(), city=payload.city.strip())
    JOBS[job_id] = job
    background.add_task(run_job, job)
    return JobStatus(
        id=job.id,
        state=job.state,
        progress=job.progress,
        message=job.message,
        started_at=job.started_at,
        finished_at=job.finished_at,
        error=job.error,
        results_count=len(job.results),
    )


@app.get("/jobs/{job_id}", response_model=JobStatus)
async def get_job(job_id: str) -> JobStatus:
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return JobStatus(
        id=job.id,
        state=job.state,
        progress=job.progress,
        message=job.message,
        started_at=job.started_at,
        finished_at=job.finished_at,
        error=job.error,
        results_count=len(job.results),
    )


@app.get("/jobs/{job_id}/events")
async def job_events(job_id: str) -> StreamingResponse:
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    async def gen() -> AsyncGenerator[bytes, None]:
        yield b"retry: 1000\n\n"
        init_payload = json.dumps(
            {
                "progress": job.progress,
                "message": job.message,
                "state": job.state,
                "error": job.error,
                "results_count": len(job.results),
            },
            ensure_ascii=False,
        )
        yield f"data: {init_payload}\n\n".encode("utf-8")
        while True:
            try:
                event = await asyncio.wait_for(job.events.get(), timeout=10.0)
            except asyncio.TimeoutError:
                yield b": keep-alive\n\n"
                continue
            payload = json.dumps(event, ensure_ascii=False)
            yield f"data: {payload}\n\n".encode("utf-8")
            if job.state in {"done", "error"} and job.progress >= 100:
                break

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/jobs/{job_id}/results", response_model=List[LeadRow])
async def get_results(job_id: str) -> List[LeadRow]:
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job.results


@app.get("/jobs/{job_id}/export.xlsx")
async def export_xlsx(job_id: str) -> Response:
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl non installato")

    wb = Workbook()
    ws = wb.active
    ws.title = "SOLAR_LEADS"

    headers = [
        "NOME AZIENDA",
        "CELLULARE",
        "INDIRIZZO",
        "CATEGORIA",
        "EMAIL",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1b2a22")
    header_font = Font(bold=True, color="36FF89")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for r in job.results:
        # already filtered to have phone
        ws.append(
            [
                r.business_name,
                r.phone,
                r.address,
                r.category,
                r.email,
            ]
        )

    # widths
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"solar_leads_{job.city}_{job.category}".replace(" ", "_") + ".xlsx"

    return Response(
        content=out.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


_INDEX_HTML = """<!doctype html>
<html lang=\"it\">
<head>
  <meta charset=\"utf-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
  <title>__APP_TITLE__</title>
  <style>
    :root {{
      --bg: #070A0B;
      --panel: rgba(255,255,255,0.04);
      --border: rgba(255,255,255,0.10);
      --text: #E6F2EF;
      --muted: rgba(230,242,239,0.65);
      --neon: #36FF89;
      --neon2: #00D46A;
    }}
    html, body {{ height: 100%; background: radial-gradient(1200px 600px at 50% 0%, rgba(54,255,137,0.07), transparent 60%), var(--bg); color: var(--text); margin: 0; font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial; }}
    .wrap {{ max-width: 1100px; margin: 0 auto; padding: 24px 18px 60px; }}
    .brand {{ text-align:center; margin-top: 10px; }}
    .brand h1 {{ margin: 10px 0 0; font-size: 42px; letter-spacing: -0.02em; }}
    .brand .sub {{ margin-top: 8px; color: var(--muted); font-size: 14px; }}
    .panel {{ margin-top: 26px; background: var(--panel); border: 1px solid var(--border); border-radius: 16px; padding: 18px; backdrop-filter: blur(10px); }}
    .row {{ display: grid; grid-template-columns: 1fr 1fr auto; gap: 12px; align-items: end; }}
    label {{ display:block; font-size: 12px; color: var(--muted); margin-bottom: 8px; }}
    input {{ width: 100%; height: 52px; border-radius: 14px; border: 1px solid rgba(54,255,137,0.25); background: rgba(0,0,0,0.25); color: var(--text); padding: 0 14px; outline: none; }}
    input:focus {{ box-shadow: 0 0 0 3px rgba(54,255,137,0.15), 0 0 18px rgba(54,255,137,0.18); border-color: rgba(54,255,137,0.6); }}
    button {{ height: 52px; border-radius: 14px; border: 1px solid rgba(54,255,137,0.35); background: linear-gradient(180deg, rgba(54,255,137,0.22), rgba(0,0,0,0.10)); color: var(--text); padding: 0 16px; cursor: pointer; font-weight: 800; }}
    button:hover {{ opacity: 0.95; box-shadow: 0 0 20px rgba(54,255,137,0.10); }}
    .presets {{ display:flex; gap: 10px; flex-wrap: wrap; margin-top: 12px; }}
    .preset {{ height: 42px; font-size: 13px; }}
    .status {{ margin-top: 16px; display:flex; justify-content: space-between; gap: 12px; color: var(--muted); font-size: 13px; }}
    .bar {{ height: 10px; background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.08); border-radius: 999px; overflow: hidden; margin-top: 10px; }}
    .bar > div {{ height: 100%; width: 0%; background: linear-gradient(90deg, var(--neon2), var(--neon)); box-shadow: 0 0 18px rgba(54,255,137,0.22); }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 18px; font-size: 13px; }}
    th, td {{ padding: 10px 10px; border-bottom: 1px solid rgba(255,255,255,0.08); vertical-align: top; }}
    th {{ text-align:left; font-size: 11px; color: rgba(230,242,239,0.70); letter-spacing: 0.18em; text-transform: uppercase; }}
    .pill {{ display:inline-block; padding: 2px 8px; border-radius: 999px; border: 1px solid rgba(54,255,137,0.35); background: rgba(54,255,137,0.10); color: var(--neon); font-weight: 800; font-size: 11px; }}
    .actions {{ display:flex; gap: 10px; align-items:center; margin-top: 12px; }}
    a {{ color: var(--neon); text-decoration: none; font-weight: 700; }}
    a:hover {{ text-decoration: underline; }}
    @media (max-width: 860px) {{
      .row {{ grid-template-columns: 1fr; }}
      button {{ width: 100%; }}
    }}
  </style>
</head>
<body>
  <div class=\"wrap\">
    <div class=\"brand\">
      <div class=\"pill\">SOLAR SNIPER</div>
      <h1>SOLAR SNIPER</h1>
      <div class=\"sub\">Scheda completa: stima consumi/potenziale FV • Link satellitare • Telefono/Email</div>
    </div>

    <div class=\"panel\">
      <div class=\"row\">
        <div>
          <label>Categoria / Target</label>
          <input id=\"category\" placeholder=\"Es: Capannoni, industrie, hotel, ristoranti...\" />
        </div>
        <div>
          <label>Città</label>
          <input id=\"city\" placeholder=\"Es: Milano\" value=\"Milano\" />
        </div>
        <div>
          <button id=\"start\">AVVIA SOLAR SNIPER</button>
        </div>
      </div>

      <div class=\"presets\">
        <button class=\"preset\" data-value=\"Fabbrica, Produzione, Industria\">🏭 FABBRICHE</button>
        <button class=\"preset\" data-value=\"Hotel, Albergo, Resort\">🏨 HOTEL</button>
        <button class=\"preset\" data-value=\"Ristorante, Pizzeria\">🍽️ RISTORANTI</button>
      </div>

      <div class=\"status\">
        <div id=\"msg\">—</div>
        <div id=\"pct\">0%</div>
      </div>
      <div class=\"bar\"><div id=\"bar\"></div></div>

      <div class=\"actions\" id=\"download\" style=\"display:none\">
        <a id=\"xlsx\" href=\"#\" target=\"_blank\" rel=\"noreferrer\">Scarica risultati (.xlsx)</a>
      </div>

      <div id=\"results\"></div>
    </div>
  </div>

<script>
  const $ = (id) => document.getElementById(id);
  const setProgress = (p, m) => {
    $('pct').textContent = (p || 0) + '%';
    $('bar').style.width = (p || 0) + '%';
    $('msg').textContent = m || '—';
  };

  document.querySelectorAll('button.preset').forEach((b) => {
    b.addEventListener('click', () => {
      $('category').value = b.getAttribute('data-value') || '';
      $('category').focus();
    });
  });

  const render = (rows) => {
    if (!rows || !rows.length) {
      $('results').innerHTML = "<div style='margin-top:16px;color:rgba(230,242,239,0.65)'>Nessun risultato (o nessun contatto trovato).</div>";
      return;
    }
    const trs = rows.map((r) => {
      const name = (r.business_name || '').replace(/</g, '&lt;');
      const phone = (r.phone || '').replace(/</g, '&lt;');
      const addr = (r.address || '').replace(/</g, '&lt;');
      const cat = (r.category || '').replace(/</g, '&lt;');
      const email = (r.email || '').replace(/</g, '&lt;');
      return `<tr><td>${name}</td><td>${phone}</td><td>${addr}</td><td>${cat}</td><td>${email}</td></tr>`;
    }).join('');
    $('results').innerHTML = `<table><thead><tr><th>Nome azienda</th><th>Cellulare / Telefono</th><th>Indirizzo</th><th>Categoria</th><th>Email</th></tr></thead><tbody>${trs}</tbody></table>`;
  };

  let jobId = null;

  const start = async () => {
    const category = ($('category').value || '').trim();
    const city = ($('city').value || '').trim();
    if (category.length < 2 || city.length < 2) {
      setProgress(0, 'Inserisci categoria e città');
      return;
    }

    $('download').style.display = 'none';
    $('results').innerHTML = '';
    setProgress(2, 'Avvio...');

    const r = await fetch('/jobs', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ category, city })
    });
    if (!r.ok) {
      setProgress(0, 'Errore avvio job');
      return;
    }
    const j = await r.json();
    jobId = j.id;

    const ev = new EventSource(`/jobs/${jobId}/events`);
    ev.onmessage = async (e) => {
      let p = null;
      try { p = JSON.parse(e.data); } catch { p = null; }
      if (!p) return;
      setProgress(p.progress || 0, p.message || '—');

      if (p.state === 'done') {
        ev.close();
        const rr = await fetch(`/jobs/${jobId}/results`);
        const rows = await rr.json();
        render(rows);
        $('xlsx').setAttribute('href', `/jobs/${jobId}/export.xlsx`);
        $('download').style.display = 'flex';
      }
      if (p.state === 'error') {
        ev.close();
        setProgress(100, p.error ? ('Errore: ' + p.error) : 'Errore');
      }
    };
    ev.onerror = () => {
      ev.close();
      setProgress(0, 'Connessione persa (SSE). Riprova.');
    };
  };

  $('start').addEventListener('click', start);
</script>
</body>
</html>"""


if __name__ == "__main__":
    import uvicorn

    print(f"{APP_TITLE} running on http://127.0.0.1:8010")
    uvicorn.run(app, host="127.0.0.1", port=8010, log_level="info")
